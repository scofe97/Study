from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter
import Data_novel
import datetime
from openpyxl.styles import Alignment
from scipy.stats import rankdata

# 엑셀 생성 , 데이터 삽입

# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.

yesterday_filename = r'날짜별 데이터\2021-12-04 노벨피아.xlsx'
today_filename = r'날짜별 데이터\2021-12-15 노벨피아.xlsx'

yesterday_wb = load_workbook(yesterday_filename, data_only=True)
today_wb = load_workbook(today_filename, data_only=True)

write_ws = today_wb['전일대비 데이터']
yesterday_ws = yesterday_wb['노벨피아 공모전']
today_ws = today_wb['노벨피아 공모전']

write_ws.sheet_properties.tabColor = "ff66ff"

today_Novels = []
yesterday_Novels = []

for row in range(2, 52):

    yesterday_Novels.append(
        Data_novel.Novel(
            yesterday_ws.cell(row, 2).value,
            yesterday_ws.cell(row, 3).value,
            int(yesterday_ws.cell(row, 4).value),
            int(yesterday_ws.cell(row, 5).value),
            int(yesterday_ws.cell(row, 6).value),
            yesterday_ws.cell(row, 1).value,
            int(yesterday_ws.cell(row,7).value)
        )
    )

    today_Novels.append(
        Data_novel.Novel(
            today_ws.cell(row, 2).value,
            today_ws.cell(row, 3).value,
            int(today_ws.cell(row, 4).value),
            int(today_ws.cell(row, 5).value),
            int(today_ws.cell(row, 6).value),
            today_ws.cell(row, 1).value,
            int(today_ws.cell(row,7).value)
        )
    )


yesterday_Novels = sorted(yesterday_Novels)
today_Novels = sorted(today_Novels)

# 출력 엑셀생성


# 조회수,  추천수,  선작수
yesterday_views_rank =[]     
today_view_rank = []

yesterday_likes_rank = []    
today_likes_rank = []

yesterday_bookmarks_rank = []     
today_bookmarks_rank = [] 


for item in yesterday_Novels:
    yesterday_views_rank.append(item.views)
    yesterday_likes_rank.append(item.likes)
    yesterday_bookmarks_rank.append(item.bookmarks)
    
for item in today_Novels:
    today_view_rank.append(item.views)
    today_likes_rank.append(item.likes)
    today_bookmarks_rank.append(item.bookmarks)
    

yesterday_views_rank = rankdata(yesterday_views_rank)
yesterday_likes_rank = rankdata(yesterday_likes_rank)
yesterday_bookmarks_rank = rankdata(yesterday_bookmarks_rank)

today_views_rank = rankdata(today_view_rank)
today_likes_rank = rankdata(today_likes_rank)
today_bookmarks_rank = rankdata(today_bookmarks_rank)


for x in range(1, len(today_Novels)+2):
    index = x-2

    if(x == 1):
        write_ws.cell(x, 1, "순위")
        write_ws.cell(x, 2, "변동")
        write_ws.cell(x, 3, "소설제목")
        write_ws.cell(x, 4, "조회수")
        write_ws.cell(x, 5, "전일대비")
        write_ws.cell(x, 6, "회차수")
        
        write_ws.cell(x, 8, "순위")
        write_ws.cell(x, 9, "변동")
        write_ws.cell(x, 10, "소설제목")
        write_ws.cell(x, 11, "추천수")
        write_ws.cell(x, 12, "전일대비")
        
        write_ws.cell(x, 14, "순위")
        write_ws.cell(x, 15, "변동")
        write_ws.cell(x, 16, "소설제목")
        write_ws.cell(x, 17, "선작수")
        write_ws.cell(x, 18, "전일대비")
        
    else:       
        write_ws.cell(x, 1, str(int(51-today_views_rank[index])) + "위")
        write_ws.cell(x, 2, today_views_rank[index] - yesterday_views_rank[index])
        write_ws.cell(x, 3, today_Novels[index].title)
        write_ws.cell(x, 4, today_Novels[index].views)
        write_ws.cell(x, 5, today_Novels[index].views-yesterday_Novels[index].views)
        write_ws.cell(x, 6, today_Novels[index].episodes)
        
        write_ws.cell(x, 8, str(int(51-today_likes_rank[index])) + "위")
        write_ws.cell(x, 9, today_likes_rank[index] - yesterday_likes_rank[index])
        write_ws.cell(x, 10, today_Novels[index].title)
        write_ws.cell(x, 11, today_Novels[index].likes)
        write_ws.cell(x, 12, today_Novels[index].likes-yesterday_Novels[index].likes)
        
        write_ws.cell(x, 14, str(int(51-today_bookmarks_rank[index])) + "위")
        write_ws.cell(x, 15, today_bookmarks_rank[index] - yesterday_bookmarks_rank[index])
        write_ws.cell(x, 16, today_Novels[index].title)
        write_ws.cell(x, 17, today_Novels[index].bookmarks)
        write_ws.cell(x, 18, today_Novels[index].bookmarks-yesterday_Novels[index].bookmarks)
        
        

# 시트 디자인

top_font = Font(name='나눔고딕', size=12, bold=True)
default_font = Font(name='나눔고딕', size=10)


for column in range(1, 19):
    
    _column = get_column_letter(column)
    
    # 정렬과 높이 맞추기
    for cell in range(len(write_ws[get_column_letter(column)])):

        if(cell == 0):
            write_ws[_column+str(cell+1)].font = top_font

            write_ws.row_dimensions[cell+1].height = 48
            
        else:
            write_ws[_column+str(cell+1)].font = default_font

            write_ws.row_dimensions[cell+1].height = 40
            
        write_ws[_column+str(cell+1)].alignment = Alignment(horizontal='center', vertical='center')


    # 글자 양식 맞추기
    if column in [2,5,9,12,15,18]:
        
        for rng in write_ws[_column+str(2):_column+str(51)]:
            for cell in rng:
                cell.number_format = '[RED]"▲"#,##0;[BLUE]"▼"#,##0;-' 
                
    if column in [4,11,17]:
        
        for rng in write_ws[_column+str(2):_column+str(51)]:
            for cell in rng:
                cell.number_format = '#,##0' 
                
                
    # 너비 맞추기   
    write_ws.column_dimensions[_column].width = 10
    
    if column in [1,2,8,9,14,15]:
        write_ws.column_dimensions[_column].width = 6
        
    if column in [3,10,16]:
        write_ws.column_dimensions[_column].width = 32
    

today_wb.save(today_filename)

