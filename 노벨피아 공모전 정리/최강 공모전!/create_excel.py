from os import name
from time import time
from openpyxl import Workbook
from scipy.stats import rankdata
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import search_novel
import datetime


Novels = search_novel.novel_information()

# 엑셀 생성 , 데이터 삽입


write_wb = Workbook()
write_ws = write_wb.create_sheet(title='노벨피아 공모전')
write_wb.create_sheet(title='전일대비 데이터')
write_wb.create_sheet(title='전주대비 데이터')
write_ws.sheet_properties.tabColor = "ff66ff"

temp = []

for item in Novels:
    temp.append(item.views)

ranks = rankdata(temp)

for x in range(1, len(Novels)+2):
    index = x-2

    if(x == 1):
        write_ws.cell(x, 1, "순위")
        write_ws.cell(x, 2, "소설제목")
        write_ws.cell(x, 3, "작가이름")
        write_ws.cell(x, 4, "조회수")
        write_ws.cell(x, 5, "회차수")
        write_ws.cell(x, 6, "좋아요")
        write_ws.cell(x, 7, "선작수")
    else:
        write_ws.cell(x, 1, str(int(51-ranks[index])) + "위")
        write_ws.cell(x, 2, Novels[index].title)
        write_ws.cell(x, 3, Novels[index].author)
        write_ws.cell(x, 4, Novels[index].views)
        write_ws.cell(x, 5, Novels[index].episodes)
        write_ws.cell(x, 6, Novels[index].likes)
        write_ws.cell(x, 7, Novels[index].bookmarks)


# 너비, 높이 조절, 정렬

top_font = Font(name='나눔고딕', size=14, bold=True)
default_font = Font(name='나눔고딕', size=10)


for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:

    for cell in range(len(write_ws[column])):
        if(cell == 0):
            write_ws[column+str(cell+1)].font = top_font

            write_ws.row_dimensions[cell+1].height = 48
        else:
            write_ws[column+str(cell+1)].font = default_font

            write_ws.row_dimensions[cell+1].height = 40
        
        write_ws[column+str(cell+1)].alignment = Alignment(horizontal='center', vertical='center')

    if column in ['D','E','F','G']:
        print(column)
        for rng in write_ws[column+str(2):column+str(51)]:
            for cell in rng:
                cell.number_format = '#,##0' 
                


write_ws.column_dimensions['A'].width = 8
write_ws.column_dimensions['B'].width = 32
write_ws.column_dimensions['C'].width = 12
write_ws.column_dimensions['D'].width = 10
write_ws.column_dimensions['E'].width = 8
write_ws.column_dimensions['F'].width = 8
write_ws.column_dimensions['G'].width = 8

# 파일생성
now = datetime.datetime.now()
nowDate = now.strftime('%Y-%m-%d')
excel_name = r'날짜별 데이터\{nowDate} 노벨피아.xlsx.'.format(nowDate=nowDate)

test_name = r'날짜별 데이터\{nowDate} test.xlsx.'.format(nowDate=nowDate)

write_wb.save(excel_name)
