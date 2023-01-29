from numpy import empty
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from 단계1_데이터_추출 import extract
from 단계1_데이터_추출2 import extract2
from 단계1_데이터_추출3 import extract3

import string
from scipy.stats import rankdata

# ? 이미지삽입 참고
# * https://blog.naver.com/PostView.naver?blogId=isc0304&logNo=222362849351&parentCategoryNo=&categoryNo=&viewDate=&isShowPopularPosts=false&from=postView

# * https://stackoverflow.com/questions/42875353/insert-an-image-from-url-in-openpyxl

# ? get_column_letter 설명
# * https://bokyeong-kim.github.io/python/basic/2020/06/14/python-basic(2).html

# ? google spreadsheet sync 사용법
# * https://carrotdesign.tistory.com/entry/Figma%EC%97%90%EC%84%9C-Data-%EC%82%AC%EC%9A%A9%ED%95%98%EA%B8%B0-Google-Sheet-Sync


# ------------------------------------------------------
# 데이터 삽입하는 코드

def create_all_excel(today_fileName,yesterday_fileName,sheetName):
        

    today_Novels = sorted(extract2(today_fileName))
    yesterday_Novels = sorted(extract2(yesterday_fileName))



    # ? 파일가져오는 코드
    dir = r'날짜별_엑셀_데이터\{0}'.format(today_fileName)
    today_wb = load_workbook(dir, data_only=True)
    create_ws = today_wb.create_sheet(sheetName)



    today_sum_views = 0
    today_sum_episodes = 0
    today_sum_likes = 0
    today_15_novels = 0
    yesterday_15_novels = 0
    tag_counter = {}

    for index in range(len(today_Novels)):
        today_sum_views += today_Novels[index].views
        today_sum_episodes += today_Novels[index].episodes
        today_sum_likes += today_Novels[index].likes
        today_15_novels += 1 if today_Novels[index].episodes >= 15 else 0
        
        if index < 100:
            for value in today_Novels[index].tag:
                try: tag_counter[value] += 1
                except: tag_counter[value ] = 1

    yesterday_sum_views = 0
    yesterday_sum_episodes = 0
    yesterday_sum_likes = 0

    for index in range(len(yesterday_Novels)):
        yesterday_sum_views += yesterday_Novels[index].views
        yesterday_sum_episodes += yesterday_Novels[index].episodes
        yesterday_sum_likes += yesterday_Novels[index].likes
        yesterday_15_novels += 1 if yesterday_Novels[index].episodes >= 15 else 0



    # ? 각 헤더별 내용
    # * 1 : 랭크
    # * 2 : 제목
    # * 3 : 19금 여부
    # * 4 : 이미지
    # * 5 : 작가
    # * 6 : 조회수 
    # * 7 : 회차수
    # * 8 : 추천수
    # * 9 : 태그

    # * 10 : 랭킹 전일대비
    # * 11 : 조회수 전일대비
    # * 12 : 랭크업
    # * 13 : 랭크다운
    # * 14 : 랭크 그룹

    for row in range(1, len(today_Novels)+2):
        
        index = row-2
        
        if row == 1:
            create_ws.cell(row,1,"rank")
            create_ws.cell(row,2,"title")
            create_ws.cell(row,3,"b_19")
            create_ws.cell(row,4,"thumbnail")
            create_ws.cell(row,5,"author")
            create_ws.cell(row,6,"views")
            create_ws.cell(row,7,"episodes")
            create_ws.cell(row,8,"likes")
            create_ws.cell(row,9,"tag")
            create_ws.cell(row,10,"view_rate")
            
            create_ws.cell(row,11,"day_to_rank")
            create_ws.cell(row,12,"day_to_views")
            create_ws.cell(row,13,"rank_up")
            create_ws.cell(row,14,"rank_down")
            create_ws.cell(row,15,"rank_view")
            create_ws.cell(row,16,"view_rate_text")
            
            create_ws.cell(row,18,"sum_views")
            create_ws.cell(row,19,"day_sum_views")
            create_ws.cell(row,20,"sum_episodes")
            create_ws.cell(row,21,"day_sum_episodes")
            create_ws.cell(row,22,"sum_likes")
            create_ws.cell(row,23,"day_sum_likes")
            
            create_ws.cell(row,25,"today_all_novels_count")
            create_ws.cell(row,26,"day_all_novels_count")
            create_ws.cell(row,27,"today_15_novels_count")
            create_ws.cell(row,28,"day_15_novels_count")
            
            create_ws.cell(row,30,"tag_name")
            create_ws.cell(row,31,"tag_count")

        else:
            
            if row == 2:
                create_ws.cell(row,18,today_sum_views)
                create_ws.cell(row,19,today_sum_views-yesterday_sum_views)
                create_ws.cell(row,20,today_sum_episodes)
                create_ws.cell(row,21,today_sum_episodes-yesterday_sum_episodes)
                create_ws.cell(row,22,today_sum_likes)
                create_ws.cell(row,23,today_sum_likes-yesterday_sum_likes)
                
                create_ws.cell(row,25,len(today_Novels))
                create_ws.cell(row,26,len(today_Novels) - len(yesterday_Novels))             
                create_ws.cell(row,27,today_15_novels)
                create_ws.cell(row,28,today_15_novels-yesterday_15_novels)
                                
            
            
            create_ws.cell(row,1,today_Novels[index].rank)
            
            # * 글자길이 24자 제한 (초과시 이후는 ...)
            if len(today_Novels[index].title) > 24:
                create_ws.cell(row,2,today_Novels[index].title[:23]+'...')
            else:               
                create_ws.cell(row,2,today_Novels[index].title)
                
            create_ws.cell(row,3,today_Novels[index].b_19)
            create_ws.cell(row,4,today_Novels[index].thumbnail)
            create_ws.cell(row,5,today_Novels[index].author)
            create_ws.cell(row,6,today_Novels[index].views)
            create_ws.cell(row,7,today_Novels[index].episodes)
            create_ws.cell(row,8,today_Novels[index].likes)
            create_ws.cell(row,9,' '.join(today_Novels[index].tag)) # * ' '.join -> List를 공백으로 구분해서 문자열로 바꿈
            
            
            if today_Novels[index].episodes > 15:    
                start = today_Novels[index].episode_start_30[4].episode_views
                end = today_Novels[index].episode_end_30[4].episode_views
                
                if start != 0 and end != 0:
                    create_ws.cell(row,10,(end/start)*100)
                    create_ws.cell(row,16,"/show")
                else:
                    create_ws.cell(row,10,0)
                    create_ws.cell(row,16,"/hide")
            else:
                create_ws.cell(row,10,0)
                create_ws.cell(row,16,"/hide")
            
            if today_Novels[index] in yesterday_Novels:
                for yesterday_Novel in yesterday_Novels:
                    if today_Novels[index] == yesterday_Novel:
                        change_rank = today_Novels[index].rank - yesterday_Novel.rank
                        create_ws.cell(row,11,abs(change_rank))
                        create_ws.cell(row,12,today_Novels[index].views - yesterday_Novel.views)
                        
                        if change_rank == 0:  
                            create_ws.cell(row,13,"/hide")
                            create_ws.cell(row,14,"/hide")
                            create_ws.cell(row,15,"/hide")
                        elif change_rank > 0:  # 랭크하락
                            create_ws.cell(row,13,"/hide")
                            create_ws.cell(row,14,"/show")       
                            create_ws.cell(row,15,"/show")       
                        elif change_rank < 0:  # 랭크상승
                            create_ws.cell(row,13,"/show")
                            create_ws.cell(row,14,"/hide")
                            create_ws.cell(row,15,"/show")
                            
            else :
                create_ws.cell(row,11,"NEW")
                create_ws.cell(row,12,today_Novels[index].views)
                create_ws.cell(row,13,"/hide")
                create_ws.cell(row,14,"/hide")
                create_ws.cell(row,15,"/hide")
                
    sorted_tag_counter = sorted(tag_counter.items(), key = lambda item: item[1], reverse = True)
    for index in range(len(sorted_tag_counter)):
        create_ws.cell(index+2,30,sorted_tag_counter[index][0])
        create_ws.cell(index+2,31,sorted_tag_counter[index][1])            
    # ------------------------------------------------------

    # ------------------------------------------------------
    # 엑셀 서식코드

    for header in range(1, 31):
        header_column = get_column_letter(header)
        
        for column in create_ws[header_column+str(2):header_column+str(len(today_Novels)+2)]:            
            # * 랭킹 서식
            for cell in column :
                if header == 1:
                    cell.number_format = '##0'

        # * 소설 데이터수치 서식                   
                if header in [6,8,18,20,22,25,27]:
                    cell.number_format = '[>=1000000]#.0#,,"M";[>=1000]#.0#,"K";##0'
                    
                elif header in [7]:
                    cell.number_format = '[>=1000000]"EP."#.0#,,"M";[>=1000]"EP."#.0#,"K";"EP."0'
                    
                elif header in [10]:
                    cell.number_format = '[==0]"/hide";#0.0#"%"'
                
                elif header in [11]:
                    cell.number_format = '[==0]"-"'
                    
                elif header in [12,19,21,23,26,28]:
                    cell.number_format = '#,##0'
                    
            
    # * 너비자동 조절
    # for column_cells  in create_ws.columns:
    #     length = max(len(str(cell.value)) for cell in column_cells)
    #     create_ws.column_dimensions[column_cells[0].column_letter].width = length+5



    # ------------------------------------------------------
    # 엑셀 저장
    today_wb.save(dir)
