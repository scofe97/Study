from numpy import empty
from openpyxl import workbook
from openpyxl import load_workbook
from 소설_양식_코드 import *
from 소설_회차 import *
import re
import string
from scipy.stats import rankdata


# ------------------------------------------------------
# 데이터 뽑아내는 코드

def extract2(filename):

    Novels = []

    # ? 파일가져오는 코드
    dir = r'날짜별_엑셀_데이터\{0}'.format(filename)
    today_wb = load_workbook(dir, data_only=True)

    if "Sheet1" in today_wb.get_sheet_names() :
        extract_ws = today_wb['Sheet1']
        extract_ws.title = "extract"
    else :
        extract_ws = today_wb['extract']


    # ? xlsx 추출하는과정
    # * 0 : 제목, (취향, 북마크, 알림) , 작가, ISBN, 19등
    # * 1 : 표지
    # * 2 : 수치, 요일, 태그
    # * 3 : 소개문구
    # * 4 : 시작부터 30화
    # * 5 : 끝부터 30화

    for row in extract_ws.iter_rows(min_row=2):
        print(row[0])
        novel_data = list(filter(None,list(map(remove_blank,row[0].value.replace('\t', "").replace('\xa0', "").splitlines()))))
        novel_b_19 = "/Show" if "19" in novel_data else "/Hide"
        novel_VEL = re.findall(r'\d+.*회',row[1].value)[0].replace("\xa0","").split()
        
        novel_VEL[0] = int(novel_VEL[0].replace("명","").replace(",",""))
        novel_VEL[1] = novel_VEL[1].replace("회차","")
        if novel_VEL[1] == "":
            novel_VEL[1] = 0
        else :
            novel_VEL[1] = int(novel_VEL[1].replace(",",""))
        novel_VEL[2] = int(novel_VEL[2].replace("회","").replace(",",""))
        
        
        novel_tag = re.findall(r'#\w+',row[1].value)
        novel_thumbnail = row[2].value
        novel_introduction = row[3].value
        
        
        novel_start_30 = []
        novel_end_30 = []
        if len(row) > 4:
            if row[4].value != '' and row[4].value != None:
                novel_start_episodes = row[4].value.replace("\t","").replace("\n","").replace("\xa0","").replace("x000D","  ").replace("x000d","  ").replace("_","").split("  ")
                for novel_episode in novel_start_episodes:         
                    
                    if novel_episode != '':
                        episode = novel_episode.split()
                        novel_start_30.append(
                            novel_episodes(
                                episode_number = -1 if episode[0] == "BONUS" else 0 if episode[0] == "Prologue" else int(episode[0].replace("EP.","")),
                                text_number =  int(episode[1].replace(",","").replace(" ","")) if len(episode) > 1 else 0,
                                episode_views = int(episode[2].replace(",","").replace(" ","")) if len(episode) > 2 else 0,
                                episode_rate = int(episode[3].replace(",","").replace(" ","")) if len(episode) == 4 else int(episode[3].replace(",","")) + int(episode[4].replace(",","").replace(" ","")) if len(episode) == 5 else 0
                            )
                        )
                
                
            if row[5].value != '' and row[5].value != None:
                novel_end_episodes = row[5].value.replace("\t","").replace("\n","").replace("\xa0","").replace("x000D","  ").replace("x000d","  ").replace("_","").split("  ")
                for novel_episode in novel_end_episodes:
                    
                    if novel_episode != '':
                        episode = novel_episode.split()                      
                        novel_end_30.append(
                            novel_episodes(
                                episode_number = -1 if episode[0] == "BONUS" else 0 if episode[0] == "Prologue" else int(episode[0].replace("EP.","")),
                                text_number =  int(episode[1].replace(",","").replace(" ","")) if len(episode) > 1 else 0,
                                episode_views = int(episode[2].replace(",","").replace(" ","")) if len(episode) > 2 else 0,
                                episode_rate = int(episode[3].replace(",","").replace(" ","")) if len(episode) == 4 else int(episode[3].replace(",","")) + int(episode[4].replace(",","").replace(" ","")) if len(episode) == 5 else 0
                            )
                        )
        
        Novels.append(
            Novel(
                title = novel_data[0],
                code = -1,
                b_19 =  novel_b_19,
                thumbnail = novel_thumbnail,
                author = novel_data[5].lstrip("작가") ,
                views = novel_VEL[0],
                episodes = novel_VEL[1],
                likes = novel_VEL[2],
                tag = novel_tag,
                episode_start_30 = novel_start_30,
                episode_end_30 = novel_end_30,
                introduction = novel_introduction
            )
        )



    # * 랭킹조회

    rankData_views = []
    # rankData_likes = []
    # rankData_bookmarks = []

    for item in Novels:
        rankData_views.append(item.views)
        # rankData_likes.append(item.likes)
        # rankData_bookmarks.append(item.likes)

    rankData_views = rankdata(rankData_views)
    # rankData_likes = rankdata(rankData_likes)
    # rankData_bookmarks = rankdata(rankData_bookmarks)   

    for index in range(len(Novels)):
        Novels[index].rank = int(len(Novels)-rankData_views[index]+1)



    return Novels


def remove_blank(str):
        return str.strip()
