from ast import IsNot
from numpy import double
from openpyxl import workbook
from openpyxl import load_workbook
from 소설_양식 import *
import string
from scipy.stats import rankdata

# ------------------------------------------------------
# 데이터 뽑아내는 코드

def extract(filename):
    
    Novels = []

    # ? 파일가져오는 코드
    dir = r'날짜별_엑셀_데이터\{0}'.format(filename)
    today_wb = load_workbook(dir, data_only=True)

    if "Sheet1" in today_wb.get_sheet_names() :
        extract_ws = today_wb['Sheet1']
        extract_ws.title = "extract"
    else :
        extract_ws = today_wb['extract']


    # ? xlsx 추출하는과정
    # * 0 : 제목
    # * 1 : 표지 URL
    # * 2 : 플러스 여부
    # * 3 : 독점 여부
    # * 4 : 19금 여부
    # * 5 : 수치 
    # * 6 : 작가 
    # * 7 : 소개글
    # * 8~15 : 태그
    # * 16 : 최근 업로드
    for row in extract_ws.iter_rows(min_row=2):
        
        # * 수치조정
        temp = row[5].value.splitlines()
        VEL = temp[0].replace(u'\xa0', u' ').split()
        
        VEL[0] = VEL[0].replace("명","")
        VEL[1] = VEL[1].replace("회차","")
        VEL[2] = VEL[2].replace("회","")
        
        for index in range(len(VEL)):
            if 'K' in VEL[index]:
                VEL[index] = float(VEL[index].replace('K',""))*1000
            elif 'M' in VEL[index]:
                VEL[index] = float(VEL[index].replace('M',""))*1000000
                
        
        # * 태그정리
        TAG = []
        for t in range(8,15):
            tag = str(row[t].value)
            if tag == "None":
                continue
            TAG.append(tag)
        Novels.append(
            Novel(
                title = row[0].value,
                b_19 =  "/Hide" if row[4].value is None else "/Show",
                thumbnail = row[1].value,
                author = row[6].value,
                views = int(VEL[0]),
                episodes = int(VEL[1]),
                likes = int(VEL[2]),
                tag = TAG,
            )
        )


    # * 랭킹조회

    rankData_views = []
    # rankData_likes = []
    # rankData_bookmarks = []

    for item in Novels:
        rankData_views.append(item.views)
        # rankData_likes.append(item.likes)
        # rankData_bookmarks.append(item.likes)

    rankData_views = rankdata(rankData_views)
    # rankData_likes = rankdata(rankData_likes)
    # rankData_bookmarks = rankdata(rankData_bookmarks)   

    for index in range(len(Novels)):
        Novels[index].rank = int(len(Novels)-rankData_views[index]+1)


    return Novels

# ------------------------------------------------------
