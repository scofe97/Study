from numpy import empty
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from 단계1_데이터_추출2 import extract2
from 단계1_데이터_추출3 import extract3

import string
from scipy.stats import rankdata

# ? 이미지삽입 참고
# * https://blog.naver.com/PostView.naver?blogId=isc0304&logNo=222362849351&parentCategoryNo=&categoryNo=&viewDate=&isShowPopularPosts=false&from=postView

# * https://stackoverflow.com/questions/42875353/insert-an-image-from-url-in-openpyxl

# ? get_column_letter 설명
# * https://bokyeong-kim.github.io/python/basic/2020/06/14/python-basic(2).html

# ? google spreadsheet sync 사용법
# * https://carrotdesign.tistory.com/entry/Figma%EC%97%90%EC%84%9C-Data-%EC%82%AC%EC%9A%A9%ED%95%98%EA%B8%B0-Google-Sheet-Sync


# ------------------------------------------------------
# 데이터 삽입하는 코드

def create_all_excel2(today_fileName,yesterday_fileName,sheetName):
        

    today_Novels = sorted(extract3(today_fileName))
    yesterday_Novels = sorted(extract3(yesterday_fileName))



    # ? 파일가져오는 코드
    dir = r'D:\study\노벨피아 공모전 정리\천지창조 공모전!\날짜별_엑셀_데이터\{0}'.format(today_fileName)
    today_wb = load_workbook(dir, data_only=True)
    create_ws = today_wb.create_sheet(sheetName)



    today_sum_views = 0
    today_sum_episodes = 0
    today_sum_likes = 0
    today_15_novels = 0
    yesterday_15_novels = 0
    tag_counter = {}

    for index in range(len(today_Novels)):
        today_sum_views += today_Novels[index].views
        today_sum_episodes += today_Novels[index].episodes
        today_sum_likes += today_Novels[index].likes
        today_15_novels += 1 if today_Novels[index].episodes >= 15 else 0
        
        if index < 100:
            for value in today_Novels[index].tag:
                try: tag_counter[value] += 1
                except: tag_counter[value ] = 1

    yesterday_sum_views = 0
    yesterday_sum_episodes = 0
    yesterday_sum_likes = 0

    for index in range(len(yesterday_Novels)):
        yesterday_sum_views += yesterday_Novels[index].views
        yesterday_sum_episodes += yesterday_Novels[index].episodes
        yesterday_sum_likes += yesterday_Novels[index].likes
        yesterday_15_novels += 1 if yesterday_Novels[index].episodes >= 15 else 0



    # ? 각 헤더별 내용
    # * 1 : 랭크
    # * 2 : 코드
    # * 3 : 제목
    # * 4 : 19금 여부
    # * 5 : 이미지
    # * 6 : 작가
    # * 7 : 조회수 
    # * 8 : 회차수
    # * 9 : 추천수
    # * 10 : 태그
    # * 11 : 연독률

    # * 12 : 랭킹 전일대비
    # * 13 : 조회수 전일대비
    # * 14 : 랭크업
    # * 15 : 랭크다운
    # * 16 : 랭크 on/off
    # * 17 : 연독률 on/off

    for row in range(1, len(today_Novels)+2):
        
        index = row-2
        
        if row == 1:
            create_ws.cell(row,1,"rank")
            create_ws.cell(row,2,"code")
            create_ws.cell(row,3,"title")
            create_ws.cell(row,4,"b_19")
            create_ws.cell(row,5,"thumbnail")
            create_ws.cell(row,6,"author")
            create_ws.cell(row,7,"views")
            create_ws.cell(row,8,"episodes")
            create_ws.cell(row,9,"likes")
            create_ws.cell(row,10,"bookmarks")
            create_ws.cell(row,11,"alarms")
            create_ws.cell(row,12,"tag")
            create_ws.cell(row,13,"view_rate")
            
            create_ws.cell(row,14,"day_to_rank")
            create_ws.cell(row,15,"day_to_views")
            create_ws.cell(row,16,"rank_up")
            create_ws.cell(row,17,"rank_down")
            create_ws.cell(row,18,"rank_view")
            create_ws.cell(row,19,"view_rate_text")
                        
            create_ws.cell(row,21,"tag_name")
            create_ws.cell(row,22,"tag_count")
            
            
            create_ws.cell(row,31,"sum_views")
            create_ws.cell(row,32,"day_sum_views")
            create_ws.cell(row,33,"sum_episodes")
            create_ws.cell(row,34,"day_sum_episodes")
            create_ws.cell(row,35,"sum_likes")
            create_ws.cell(row,36,"day_sum_likes")
            
            create_ws.cell(row,37,"today_all_novels_count")
            create_ws.cell(row,38,"day_all_novels_count")
            create_ws.cell(row,39,"today_15_novels_count")
            create_ws.cell(row,40,"day_15_novels_count")


        else:
            
            if row == 2:
                create_ws.cell(row,31,today_sum_views)
                create_ws.cell(row,32,today_sum_views-yesterday_sum_views)
                create_ws.cell(row,33,today_sum_episodes)
                create_ws.cell(row,34,today_sum_episodes-yesterday_sum_episodes)
                create_ws.cell(row,35,today_sum_likes)
                create_ws.cell(row,36,today_sum_likes-yesterday_sum_likes)
                
                create_ws.cell(row,37,len(today_Novels))
                create_ws.cell(row,38,len(today_Novels) - len(yesterday_Novels))             
                create_ws.cell(row,39,today_15_novels)
                create_ws.cell(row,40,today_15_novels-yesterday_15_novels)
                                
            
            
            create_ws.cell(row,1,today_Novels[index].rank)
            create_ws.cell(row,2,today_Novels[index].code)
                        
            # * 글자길이 24자 제한 (초과시 이후는 ...)
            if len(today_Novels[index].title) > 24:
                create_ws.cell(row,3,today_Novels[index].title[:23]+'...')
            else:               
                create_ws.cell(row,3,today_Novels[index].title)
                
            create_ws.cell(row,4,today_Novels[index].b_19)
            create_ws.cell(row,5,today_Novels[index].thumbnail)
            create_ws.cell(row,6,today_Novels[index].author)
            create_ws.cell(row,7,today_Novels[index].views)
            create_ws.cell(row,8,today_Novels[index].episodes)
            create_ws.cell(row,9,today_Novels[index].likes)
            create_ws.cell(row,10,today_Novels[index].bookmarks)
            create_ws.cell(row,11,today_Novels[index].alarms)
            create_ws.cell(row,12,' '.join(today_Novels[index].tag)) # * ' '.join -> List를 공백으로 구분해서 문자열로 바꿈
            
            if today_Novels[index].episodes > 15:    
                if len(today_Novels[index].episode_end_30) > 5:
                    start = today_Novels[index].episode_start_30[5].episode_views + today_Novels[index].episode_start_30[4].episode_views
                    end = today_Novels[index].episode_end_30[5].episode_views + today_Novels[index].episode_end_30[4].episode_views
                
                    if start != 0 and end != 0:
                        create_ws.cell(row,13,(end/start)*100)
                        create_ws.cell(row,19,"/show")
                    else:
                        create_ws.cell(row,13,0)
                        create_ws.cell(row,19,"/hide")
            else:
                create_ws.cell(row,13,0)
                create_ws.cell(row,19,"/hide")
            
            if today_Novels[index] in yesterday_Novels:
                for yesterday_Novel in yesterday_Novels:
                    if today_Novels[index] == yesterday_Novel:
                        change_rank = today_Novels[index].rank - yesterday_Novel.rank
                        create_ws.cell(row,14,abs(change_rank))
                        create_ws.cell(row,15,today_Novels[index].views - yesterday_Novel.views)
                        
                        if change_rank == 0:  
                            create_ws.cell(row,16,"/hide")
                            create_ws.cell(row,17,"/hide")
                            create_ws.cell(row,18,"/hide")
                        elif change_rank > 0:  # 랭크하락
                            create_ws.cell(row,16,"/hide")
                            create_ws.cell(row,17,"/show")       
                            create_ws.cell(row,18,"/show")       
                        elif change_rank < 0:  # 랭크상승
                            create_ws.cell(row,16,"/show")
                            create_ws.cell(row,17,"/hide")
                            create_ws.cell(row,18,"/show")
                            
            else :
                create_ws.cell(row,14,"NEW")
                create_ws.cell(row,15,today_Novels[index].views)
                create_ws.cell(row,16,"/hide")
                create_ws.cell(row,17,"/hide")
                create_ws.cell(row,18,"/hide")
                
    sorted_tag_counter = sorted(tag_counter.items(), key = lambda item: item[1], reverse = True)
    for index in range(len(sorted_tag_counter)):
        create_ws.cell(index+2,21,sorted_tag_counter[index][0])
        create_ws.cell(index+2,22,sorted_tag_counter[index][1])            
    # ------------------------------------------------------

    # ------------------------------------------------------
    # 엑셀 서식코드

    for header in range(1, 50):
        header_column = get_column_letter(header)
        
        for column in create_ws[header_column+str(2):header_column+str(len(today_Novels)+2)]:            
            # * 랭킹 서식
            for cell in column :
                if header in[1,32,34,36,38,40]:
                    cell.number_format = '#,##0'
        # * 소설 데이터수치 서식                   
                elif header in [7,9,10,11,15,31,33,35,37,39]:
                    cell.number_format = '[>=1000000]#.0#,,"M";[>=1000]#.0#,"K";##0'
                    
                elif header in [8]:
                    cell.number_format = '[>=1000000]"EP."#.0#,,"M";[>=1000]"EP."#.0#,"K";"EP."0'
                    
                elif header in [13]:
                    cell.number_format = '[==0]"/hide";#0.0#"%"'
                
                elif header in [14]:
                    cell.number_format = '[==0]"-"'
                    
                
    # * 너비자동 조절
    # for column_cells  in create_ws.columns:
    #     length = max(len(str(cell.value)) for cell in column_cells)
    #     create_ws.column_dimensions[column_cells[0].column_letter].width = length+5



    # ------------------------------------------------------
    # 엑셀 저장
    today_wb.save(dir)
